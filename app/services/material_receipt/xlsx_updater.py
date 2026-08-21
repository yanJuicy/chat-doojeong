# -*- coding: utf-8 -*-
"""
자재입출고.xlsx의 "구입" 시트를 주문서(OrderDocument) 내용으로 업데이트한다.

시트 구조 (실측 확인):
  - 1행: 제목("▲ 자재 구매 입고현황 ( 2026년 07월 )") — 여기서 시트가 몇 년/몇 월인지 읽는다.
  - 3행: 대분류 헤더(업체/라인/차종/품번/품명/... /1주/2주/... 반복 두 번:
         하나는 주간 합계 블록, 하나는 일별 상세 블록)
  - 4행: 일별 상세 블록의 요일(일/월/화/수/목/금/토)
  - 5행: 일별 상세 블록의 실제 "일(day of month)" 숫자 — 여기서 날짜→열 매핑을 계산
  - 6행부터 데이터: F열=품번, G열=품명, B열=업체
  - 각 날짜는 2개 열(수량, 금액)을 쓰며 금액은 '=단가*수량' 수식으로 자동 계산되므로
    우리가 손으로 채우는 건 "수량" 열 하나뿐이다.

이 모듈은:
  1) 시트 제목에서 연/월을 읽어, 채우려는 날짜와 월이 일치하는지 검증한다(다르면 명확히 에러).
  2) 날짜 → "수량" 열 위치를 계산한다.
  3) 품번(F열)으로 행을 찾는다 — 새 품번은 자동으로 행을 만들지 않고 목록으로 보고한다
     (수식이 많은 시트라 행 삽입을 잘못하면 다른 행 참조가 깨질 위험이 커서, 신규 품목은
     사람이 직접 확인 후 추가하는 쪽이 안전하다고 판단함).
  4) 매칭된 품번의 해당 날짜 "수량" 셀에 값을 더한다(기존 값이 있으면 누적).
"""
import re
from dataclasses import dataclass, field

import openpyxl

MONTH_TITLE_RE = re.compile(r"(\d{4})\s*년\s*0?(\d{1,2})\s*월")


@dataclass
class ApplyResult:
    matched: list = field(default_factory=list)      # [(품번, 품명, 행번호, 적용수량, 셀주소)]
    unmatched: list = field(default_factory=list)     # [(품번, 품명, 사유)]
    sheet_year: int = 0
    sheet_month: int = 0


def _find_sheet_period(ws) -> tuple[int, int]:
    """1행 제목에서 '2026년 07월' 형태를 찾아 (년, 월) 반환."""
    for row in ws.iter_rows(min_row=1, max_row=2, max_col=20):
        for cell in row:
            if isinstance(cell.value, str):
                m = MONTH_TITLE_RE.search(cell.value)
                if m:
                    return int(m.group(1)), int(m.group(2))
    raise ValueError("시트 제목에서 '2026년 07월' 형태의 연/월을 찾지 못했습니다.")


def _find_day_columns(ws, header_row: int = 5, search_max_col: int = 135) -> dict:
    """5행(일자 숫자)을 훑어 {day_of_month: 수량컬럼} 매핑을 만든다.
    같은 day 숫자가 여러 번 나올 수 있어(월 경계, 미사용 6주차 placeholder 등),
    '처음 나오는 1'부터 연속으로 증가하는 첫 구간만 신뢰한다.
    """
    raw = []
    for c in range(1, search_max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, (int, float)) and v and v == int(v):
            raw.append((c, int(v)))

    # '1'이 처음 등장하는 지점부터, 값이 1씩 증가하는 연속 구간만 채택
    start_idx = next((i for i, (_, d) in enumerate(raw) if d == 1), None)
    if start_idx is None:
        raise ValueError("일자 헤더(5행)에서 '1일' 시작 위치를 찾지 못했습니다.")

    day_to_col = {}
    expected = 1
    for c, d in raw[start_idx:]:
        if d != expected:
            break
        day_to_col[d] = c
        expected += 1
    return day_to_col


def _find_item_row(ws, item_code: str, code_col: int = 6, min_row: int = 8, max_row: int = 500):
    for r in range(min_row, max_row + 1):
        v = ws.cell(row=r, column=code_col).value
        if v and str(v).strip() == item_code.strip():
            return r
    return None


def apply_order_to_workbook(wb, order, sheet_name: str = "구입") -> ApplyResult:
    """order: order_parser.OrderDocument. wb는 openpyxl Workbook(수식 보존을 위해 data_only=False로 로드된 것)."""
    ws = wb[sheet_name]
    year, month = _find_sheet_period(ws)

    result = ApplyResult(sheet_year=year, sheet_month=month)

    if order.written_at is None:
        raise ValueError("주문서에서 작성일을 읽지 못해 어느 날짜 칸에 넣을지 알 수 없습니다.")

    if order.written_at.year != year or order.written_at.month != month:
        raise ValueError(
            f"이 시트는 {year}년 {month}월용인데, 주문서 작성일은 "
            f"{order.written_at.year}년 {order.written_at.month}월입니다 — "
            f"{order.written_at.year}년 {order.written_at.month}월 시트를 열어서 다시 실행하세요."
        )

    day_to_col = _find_day_columns(ws)
    day = order.written_at.day
    if day not in day_to_col:
        raise ValueError(f"{month}월 {day}일에 해당하는 열을 시트에서 찾지 못했습니다.")
    qty_col = day_to_col[day]

    for item in order.items:
        row = _find_item_row(ws, item.item_code)
        if row is None:
            result.unmatched.append((item.item_code, item.item_name, "시트에 등록되지 않은 품번(신규 품목으로 추정) — 수동 확인 필요"))
            continue

        cell = ws.cell(row=row, column=qty_col)
        prev = cell.value if isinstance(cell.value, (int, float)) else 0
        cell.value = prev + item.quantity

        col_letter = openpyxl.utils.get_column_letter(qty_col)
        result.matched.append((item.item_code, item.item_name, row, item.quantity, f"{col_letter}{row}"))

    return result


if __name__ == "__main__":
    import sys
    sys.path.insert(0, ".")
    from order_parser import parse_order_doc

    order_path, xlsx_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    order = parse_order_doc(order_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    result = apply_order_to_workbook(wb, order)

    print(f"시트 기준월: {result.sheet_year}년 {result.sheet_month}월")
    print(f"매칭 성공 {len(result.matched)}건:")
    for code, name, row, qty, addr in result.matched:
        print(f"  - {code} ({name}) → {addr} 에 +{qty}")
    if result.unmatched:
        print(f"매칭 실패 {len(result.unmatched)}건:")
        for code, name, reason in result.unmatched:
            print(f"  - {code} ({name}): {reason}")

    wb.save(out_path)
    print(f"\n저장 완료: {out_path}")
